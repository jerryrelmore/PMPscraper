#! python3

########################################################################
# Author:   Jerry Elmore
# Rev:      0.1a
# Date:     2-Feb-2018
# Program:  PMPscraper.py
#
# This program reads an Excel/CSV file for electronic component part numbers,
# runs those part numbers through the major authorized distribution websites,
# and scrapes pricing from each.

########################################################################

import sys
import os
import openpyxl
import urllib.request
import csv
import pprint
from datetime import datetime
from bs4 import BeautifulSoup
from urllist import distroURLs

################################################################################
#                             DEFINE ALL FUNCTIONS                             #
##### Read the part numbers and relevant info from the PMP RFQ Excel file ######
def readExcel(filename):
    try:    # Run in try/except syntax to better handle exceptions
        wb = openpyxl.load_workbook(filename)   # Assign local variable wb (workbook) the name of the Excel file passed to function from global variable excelFilename
        print(filename, "successfully opened...") # Let user know the file was successfully opened
        sheet = wb.worksheets[0] # Assign first workbook sheet name to a variable for later manipulation
        print("\nUsing data from workbook sheet:", sheet) # Let user know the data from first sheet in the workbook will be used
        partNumberData = {} # Initialize dictionary where data from the spreadsheet will be stored so it can be manipulated
        print("\nReading rows...", end='') # Let user known the data is getting ready to be read

        for row in range (2, sheet.max_row):
            # Each row in the spreadsheet has data for one part number - variables below are for the part number columns
            lineNumber          = sheet['A' + str(row)].value
            customerPartNumber  = sheet['B' + str(row)].value
            mfrPartNumber       = sheet['C' + str(row)].value
            pinPackage          = sheet['D' + str(row)].value
            mfrName             = sheet['E' + str(row)].value
            dateCodeRequirement = sheet['F' + str(row)].value
            needDate            = sheet['G' + str(row)].value
            quantity1           = sheet['H' + str(row)].value
            quantity2           = sheet['I' + str(row)].value
            quantity3           = sheet['J' + str(row)].value
            print(".", end='')

            # Make sure the key for this line number exists (essentially creating a key for a dictionary within a dictionary
            partNumberData.setdefault(lineNumber, {})
            # General data format is:
            partNumberData[lineNumber] = {'customerPartNumber': customerPartNumber,
                                          'mfrPartNumber': mfrPartNumber,
                                          'mfrName': mfrName,
                                          'dateCodeRequirement': dateCodeRequirement,
                                          'needDate': needDate,
                                          'quantity1': quantity1,
                                          'quantity2': quantity2,
                                          'quantity3': quantity3}
            
            # Make sure the key for this customer part number on this line numbe
    except Exception as e:
       print("Error:", e)
       print("\n\n\n\nAn error occurred - try running the program again")
       sys.exit() # Forcefully end the program

    print("done") # Let user know program is done reading Excel data into the dictionary

    print("\n\nDo you want to review the data? (Y/N): ", end='')
    dataReview = "!!"
    while (dataReview != "Y" and dataReview != "y" and dataReview != "N" and dataReview != "n"):
        dataReview = input()

    # If customer wants to review, the next few lines will take them through bite-size chunks of the partNumberData dictionary
    if (dataReview == "Y" or dataReview == "y"):
        for i in range (1, sheet.max_row - 1):  # Loop through the entire dictionary
            print(i, ": ", end='')              # Print the index key for the data that're displayed
            pprint.pprint(partNumberData[i])    # Pretty print the data
            print("\n")                         # Add an extra line between prints
            if (i % 5 == 0 and i != sheet.max_row): # Pause output after five indices are printed
                print("Press Enter to continue...")
                input()
    return partNumberData # Send the data back to the calling function
################################################################################

def accessDistroSite():
    # Function procedures will go here
    return
################################################################################

def accessArrow():
    # Function procedures will go here
    return
################################################################################

def accessAvnet():
    # Function procedures will go here
    return
################################################################################

def accessDigiKey():
    # Function procedures will go here
    return
################################################################################

def accessMouser():
    # Function procedures will go here
    return
################################################################################

def writeScrapedData():
    # Function procedures will go here
    return
################################################################################


# INITIALIZE
############## Initialize the program and check filename validity ##############
os.system('cls')    # clear screen

if len(sys.argv) > 1:   # make sure argument of name of file to open was passed on command line
    excelFilename = ''.join(sys.argv[1:]) # Remove first command line argument (this program's name)
    if os.path.isfile(excelFilename):
        print("Opening: ", excelFilename) # print name of RFQ form file to be parsed    
    else: # error if user passed a non-existent filename to the program
        print(excelFilename, "is not a valid file name. Try running the program again.")
        sys.exit() # End the program
else: # error if no filename was passed to the program at all
    print("Make sure you pass a file name on the command line --- syntax: python PMPscraper.py PMP_RFQ_FORM.xlsx")
    sys.exit() # End the program

ARROW_URL = distroURLs['ARROW_URL']
AVNET_URL = distroURLs['AVNET_URL']
DIGIKEY_URL = distroURLs['DIGIKEY_URL']
MOUSER_URL = distroURLs['MOUSER_URL']
################################################################################


# MAIN
############## Call the various functions as needed ############################

spreadsheetData = readExcel(excelFilename) # call the readExcel function and have it return a dictionary full of all the spreadsheet data
#pprint.pprint(spreadsheetData) # DEBUG PRINT: to make sure that the data is transferred back out of readExcel() appropriately
